"""
Excel və PDF smeta generatoru
"""

import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, Image as RLImage
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from config import COMPANY_NAME, COMPANY_PHONE, COMPANY_EMAIL, COMPANY_ADDRESS, COMPANY_LOGO, CURRENCY, OUTPUT_DIR

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Font qeydiyyatı (Azərbaycan hərfləri üçün) ────────────────────────────────
import reportlab
_RL_DIR = os.path.dirname(reportlab.__file__)

def _reg(name, filename):
    paths = [
        f"C:/Windows/Fonts/{filename}",
        os.path.join(os.path.dirname(__file__), filename),
        f"/usr/share/fonts/truetype/{filename}",
    ]
    for p in paths:
        if os.path.exists(p):
            pdfmetrics.registerFont(TTFont(name, p))
            return True
    return False

_has_font = all([
    _reg("Arial",            "arial.ttf"),
    _reg("Arial-Bold",       "arialbd.ttf"),
    _reg("Arial-Italic",     "ariali.ttf"),
    _reg("Arial-BoldItalic", "arialbi.ttf"),
])

if _has_font:
    FONT        = "Arial"
    FONT_BOLD   = "Arial-Bold"
    FONT_ITALIC = "Arial-Italic"
    FONT_BI     = "Arial-BoldItalic"
else:
    FONT        = "Helvetica"
    FONT_BOLD   = "Helvetica-Bold"
    FONT_ITALIC = "Helvetica-Oblique"
    FONT_BI     = "Helvetica-BoldOblique"

# Rənglər
DARK_BLUE  = "1B2A4A"
ACCENT     = "C9973A"   # qızılı
LIGHT_GRAY = "F5F6FA"
MID_GRAY   = "DDE1EA"
WHITE      = "FFFFFF"


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, name="Calibri", italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def generate_excel(smeta: dict) -> str:
    wb = Workbook()

    # ── Əsas vərəq ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Smeta"

    # Sütun genişlikləri
    widths = [4, 32, 12, 12, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Başlıq bloku ────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = COMPANY_NAME.upper()
    c.font = _font(bold=True, size=18, color=WHITE)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 36
    row += 1

    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = f"Tel: {COMPANY_PHONE}  |  {COMPANY_EMAIL}"
    c.font = _font(size=9, color="CCCCCC")
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 18
    row += 2

    # ── Smeta məlumatları ────────────────────────────────────────────────────
    info_pairs = [
        ("Smeta №",      smeta["smeta_number"]),
        ("Müştəri",      smeta["client_name"]),
        ("Telefon",      smeta["client_phone"]),
        ("Ünvan",        smeta["address"]),
        ("Tarix",        datetime.now().strftime("%d.%m.%Y")),
    ]
    for label, val in info_pairs:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"].value = label
        ws[f"A{row}"].font = _font(bold=True, size=10, color=DARK_BLUE)
        ws[f"A{row}"].fill = _fill(LIGHT_GRAY)
        ws.merge_cells(f"C{row}:F{row}")
        ws[f"C{row}"].value = val
        ws[f"C{row}"].font = _font(size=10)
        ws.row_dimensions[row].height = 18
        row += 1
    row += 1

    # ── Cədvəl başlığı ───────────────────────────────────────────────────────
    headers = ["№", "İşin adı", "Ölçü", "Miqdar", f"Qiymət ({CURRENCY})", f"Məbləğ ({CURRENCY})"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _font(bold=True, size=10, color=WHITE)
        c.fill = _fill(DARK_BLUE)
        c.alignment = _align("center")
        c.border = _border()
    ws.row_dimensions[row].height = 22
    row += 1

    # ── İş sətirləri ────────────────────────────────────────────────────────
    item_counter = 1
    rooms_data = smeta["rooms_data"]

    for room_name, categories in rooms_data.items():
        # Otaq başlığı
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = f"📍 {room_name}"
        c.font = _font(bold=True, size=11, color=WHITE)
        c.fill = _fill(ACCENT)
        c.alignment = _align("left")
        c.border = _border()
        ws.row_dimensions[row].height = 20
        row += 1

        for cat_name, items in categories.items():
            if not items:
                continue
            # Kateqoriya başlığı
            ws.merge_cells(f"A{row}:F{row}")
            c = ws[f"A{row}"]
            c.value = f"  {cat_name}"
            c.font = _font(bold=True, size=10, color=DARK_BLUE, italic=True)
            c.fill = _fill(MID_GRAY)
            c.border = _border()
            ws.row_dimensions[row].height = 18
            row += 1

            for item in items:
                name    = item["name"]
                unit    = item["unit"]
                qty     = item["qty"]
                price   = item["price"]
                amount  = qty * price

                vals = [item_counter, name, unit, qty, price, amount]
                for col, v in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.font = _font(size=10)
                    c.border = _border()
                    c.alignment = _align("center" if col in (1, 3, 4) else "left")
                    if col == 6:
                        c.number_format = '#,##0.00'
                    if col == 5:
                        c.number_format = '#,##0.00'
                    # alternating row color
                    if item_counter % 2 == 0:
                        c.fill = _fill("F9FAFB")
                ws.row_dimensions[row].height = 18
                row += 1
                item_counter += 1

    row += 1

    # ── Yekun hesablamalar ───────────────────────────────────────────────────
    subtotal = smeta["subtotal"]
    margin   = subtotal * smeta["margin_pct"] / 100
    after_m  = subtotal + margin
    discount = after_m * smeta["discount_pct"] / 100
    after_d  = after_m - discount
    vat      = after_d * smeta["vat_pct"] / 100
    total    = after_d + vat

    totals = [
        ("Cəmi (işlər):",         subtotal),
        (f"Marja ({smeta['margin_pct']}%):", margin),
        (f"Endirim ({smeta['discount_pct']}%):", -discount),
        (f"ƏDV ({smeta['vat_pct']}%):", vat),
    ]
    for label, val in totals:
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]
        c.value = label
        c.font = _font(bold=True, size=10, color=DARK_BLUE)
        c.fill = _fill(LIGHT_GRAY)
        c.alignment = _align("right")
        c.border = _border()
        c = ws.cell(row=row, column=6, value=val)
        c.font = _font(bold=True, size=10)
        c.number_format = '#,##0.00'
        c.border = _border()
        row += 1

    # Yekun total
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = f"YEKUNİ ({CURRENCY}):"
    c.font = _font(bold=True, size=13, color=WHITE)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align("right")
    c.border = _border()
    c = ws.cell(row=row, column=6, value=total)
    c.font = _font(bold=True, size=13, color=WHITE)
    c.fill = _fill(DARK_BLUE)
    c.number_format = '#,##0.00'
    c.border = _border()
    ws.row_dimensions[row].height = 26
    row += 2

    # ── Qeydlər ─────────────────────────────────────────────────────────────
    if smeta.get("notes"):
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"].value = f"Qeydlər: {smeta['notes']}"
        ws[f"A{row}"].font = _font(size=9, italic=True, color="666666")
        row += 1

    # ── İmza bölməsi ────────────────────────────────────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].value = "Müştərinin imzası: _______________"
    ws[f"A{row}"].font = _font(size=10)
    ws.merge_cells(f"D{row}:F{row}")
    ws[f"D{row}"].value = "İcraçının imzası: _______________"
    ws[f"D{row}"].font = _font(size=10)

    # Faylı saxla
    fname = f"{OUTPUT_DIR}/smeta_{smeta['smeta_number'].replace('-','_')}.xlsx"
    wb.save(fname)
    return fname


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF
# ═══════════════════════════════════════════════════════════════════════════════

def generate_pdf(smeta: dict) -> str:
    fname = f"{OUTPUT_DIR}/smeta_{smeta['smeta_number'].replace('-','_')}.pdf"

    doc = SimpleDocTemplate(
        fname,
        pagesize=A4,
        leftMargin=1.5*cm,
        rightMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    story  = []

    dark  = colors.HexColor(f"#{DARK_BLUE}")
    gold  = colors.HexColor(f"#{ACCENT}")
    lgray = colors.HexColor(f"#{LIGHT_GRAY}")
    mgray = colors.HexColor(f"#{MID_GRAY}")

    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    title_style = ps("Title",
        fontName=FONT_BOLD, fontSize=16,
        textColor=colors.white, alignment=TA_CENTER, leading=20)
    sub_style = ps("Sub",
        fontName=FONT, fontSize=8,
        textColor=colors.lightgrey, alignment=TA_CENTER, leading=10)
    info_label = ps("InfoLabel",
        fontName=FONT_BOLD, fontSize=9, textColor=dark)
    info_val = ps("InfoVal",
        fontName=FONT, fontSize=9, textColor=colors.black)
    room_style = ps("Room",
        fontName=FONT_BOLD, fontSize=10,
        textColor=colors.white, alignment=TA_LEFT)
    cat_style = ps("Cat",
        fontName=FONT_BI, fontSize=8,
        textColor=dark)
    body_style = ps("Body",
        fontName=FONT, fontSize=8, textColor=colors.black)
    total_style = ps("Total",
        fontName=FONT_BOLD, fontSize=12,
        textColor=colors.white, alignment=TA_RIGHT)

    # ── Başlıq ──────────────────────────────────────────────────────────────
    # Logo
    logo_cell = ""
    if os.path.exists(COMPANY_LOGO):
        logo_cell = RLImage(COMPANY_LOGO, width=4*cm, height=2*cm)

    header_data = [
        [logo_cell, Paragraph(
            f'<font size="16"><b>{COMPANY_NAME.upper()}</b></font><br/>'
            f'<font size="10" color="#C9973A">TƏMİR SMETASİ</font><br/>'
            f'<font size="8" color="#AAAAAA">Tel: {COMPANY_PHONE}  |  {COMPANY_ADDRESS}</font>',
            ps("HInfo", fontName=FONT, fontSize=10,
               textColor=colors.white, alignment=TA_LEFT, leading=16)
        )],
    ]
    header_tbl = Table(header_data, colWidths=[4.5*cm, 13.5*cm])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), dark),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 10),
        ("LEFTPADDING",  (0, 0), (0, -1), 10),
        ("LEFTPADDING",  (1, 0), (1, -1), 15),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 0.3*cm))

    # ── Smeta info ───────────────────────────────────────────────────────────
    info_rows = [
        ["Smeta №",  smeta["smeta_number"],  "Tarix", datetime.now().strftime("%d.%m.%Y")],
        ["Müştəri",  smeta["client_name"],   "Telefon", smeta["client_phone"]],
        ["Ünvan",    smeta["address"],        "", ""],
    ]
    info_tbl = Table(info_rows, colWidths=[3*cm, 6*cm, 3*cm, 6*cm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0, 0), (-1, -1), FONT),
        ("FONTSIZE",  (0, 0), (-1, -1), 9),
        ("FONTNAME",  (0, 0), (0, -1), FONT_BOLD),
        ("FONTNAME",  (2, 0), (2, -1), FONT_BOLD),
        ("BACKGROUND",(0, 0), (-1, -1), lgray),
        ("GRID",      (0, 0), (-1, -1), 0.5, mgray),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── Cədvəl başlıqları ────────────────────────────────────────────────────
    col_w = [1*cm, 7.5*cm, 2*cm, 2*cm, 2.5*cm, 3*cm]
    tbl_data = [[
        Paragraph("<b>№</b>", ps("H", fontName=FONT_BOLD, fontSize=9, textColor=colors.white, alignment=TA_CENTER)),
        Paragraph("<b>İşin adı</b>", ps("H2", fontName=FONT_BOLD, fontSize=9, textColor=colors.white)),
        Paragraph("<b>Ölçü</b>", ps("H", fontName=FONT_BOLD, fontSize=9, textColor=colors.white, alignment=TA_CENTER)),
        Paragraph("<b>Miqdar</b>", ps("H", fontName=FONT_BOLD, fontSize=9, textColor=colors.white, alignment=TA_CENTER)),
        Paragraph(f"<b>Qiymət</b>", ps("H", fontName=FONT_BOLD, fontSize=9, textColor=colors.white, alignment=TA_CENTER)),
        Paragraph(f"<b>Məbləğ</b>", ps("H", fontName=FONT_BOLD, fontSize=9, textColor=colors.white, alignment=TA_CENTER)),
    ]]

    item_counter = 1
    for room_name, categories in smeta["rooms_data"].items():
        # Otaq başlığı
        tbl_data.append([
            Paragraph(f"📍 {room_name}", room_style),
            "", "", "", "", ""
        ])

        for cat_name, items in categories.items():
            if not items:
                continue
            tbl_data.append([
                Paragraph(f"  {cat_name}", cat_style),
                "", "", "", "", ""
            ])
            for item in items:
                amt = item["qty"] * item["price"]
                tbl_data.append([
                    str(item_counter),
                    item["name"],
                    item["unit"],
                    str(item["qty"]),
                    f"{item['price']:,.2f}",
                    f"{amt:,.2f}",
                ])
                item_counter += 1

    main_tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)

    style_cmds = [
        # Başlıq sırası
        ("BACKGROUND",   (0, 0), (-1, 0), dark),
        ("TOPPADDING",   (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING",(0, 0), (-1, 0), 6),
        ("FONTNAME",     (0, 0), (-1, -1), FONT),
        ("FONTSIZE",     (0, 0), (-1, -1), 8),
        ("GRID",         (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",   (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 1), (-1, -1), 3),
        ("ALIGN",        (0, 0), (0, -1), "CENTER"),
        ("ALIGN",        (2, 0), (-1, -1), "CENTER"),
    ]
    # Rənglər üçün sıraya baxaraq
    for i, row_data in enumerate(tbl_data[1:], 1):
        v = str(row_data[0])
        if v.startswith("📍"):
            style_cmds += [
                ("BACKGROUND",  (0, i), (-1, i), gold),
                ("FONTNAME",    (0, i), (-1, i), FONT_BOLD),
                ("TEXTCOLOR",   (0, i), (-1, i), colors.white),
                ("SPAN",        (0, i), (-1, i)),
            ]
        elif v.startswith(" "):
            style_cmds += [
                ("BACKGROUND",  (0, i), (-1, i), mgray),
                ("FONTNAME",    (0, i), (-1, i), FONT_BI),
                ("SPAN",        (0, i), (-1, i)),
            ]
        elif i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), lgray))

    main_tbl.setStyle(TableStyle(style_cmds))
    story.append(main_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── Yekun hesablamalar ───────────────────────────────────────────────────
    subtotal = smeta["subtotal"]
    margin   = subtotal * smeta["margin_pct"] / 100
    after_m  = subtotal + margin
    discount = after_m * smeta["discount_pct"] / 100
    after_d  = after_m - discount
    vat      = after_d * smeta["vat_pct"] / 100
    total    = after_d + vat

    sum_rows = [
        [f"Cəmi (işlər):",                f"{subtotal:,.2f} {CURRENCY}"],
        [f"Marja ({smeta['margin_pct']}%):", f"{margin:,.2f} {CURRENCY}"],
        [f"Endirim ({smeta['discount_pct']}%):", f"-{discount:,.2f} {CURRENCY}"],
        [f"ƏDV ({smeta['vat_pct']}%):",     f"{vat:,.2f} {CURRENCY}"],
    ]
    sum_tbl = Table(sum_rows, colWidths=[14*cm, 4*cm])
    sum_tbl.setStyle(TableStyle([
        ("FONTNAME",     (0, 0), (-1, -1), FONT),
        ("FONTSIZE",     (0, 0), (-1, -1), 9),
        ("FONTNAME",     (0, 0), (0, -1), FONT_BOLD),
        ("ALIGN",        (1, 0), (1, -1), "RIGHT"),
        ("BACKGROUND",   (0, 0), (-1, -1), lgray),
        ("GRID",         (0, 0), (-1, -1), 0.3, mgray),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]))
    story.append(sum_tbl)

    total_row = [[
        Paragraph(f"YEKUNİ ({CURRENCY}):", total_style),
        Paragraph(f"<b>{total:,.2f}</b>", ps("TV", fontName=FONT_BOLD, fontSize=12, textColor=colors.white, alignment=TA_RIGHT)),
    ]]
    total_tbl = Table(total_row, colWidths=[14*cm, 4*cm])
    total_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), dark),
        ("TOPPADDING",   (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 8),
    ]))
    story.append(total_tbl)

    if smeta.get("notes"):
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph(f"<i>Qeydlər: {smeta['notes']}</i>",
            ps("N", fontName=FONT_ITALIC, fontSize=8, textColor=colors.grey)))

    # İmza
    story.append(Spacer(1, 1*cm))
    sign_tbl = Table([
        ["Müştəri: ___________________________", "İcraçı: ___________________________"]
    ], colWidths=[9*cm, 9*cm])
    sign_tbl.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), FONT),
        ("FONTSIZE", (0,0), (-1,-1), 9),
    ]))
    story.append(sign_tbl)

    doc.build(story)
    return fname


# ═══════════════════════════════════════════════════════════════════════════════
#  AYLIK HESABAT EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

MONTH_NAMES_AZ = {
    "01": "Yanvar",  "02": "Fevral",  "03": "Mart",
    "04": "Aprel",   "05": "May",     "06": "İyun",
    "07": "İyul",    "08": "Avqust",  "09": "Sentyabr",
    "10": "Oktyabr", "11": "Noyabr",  "12": "Dekabr",
}


def generate_monthly_excel(report: dict, smetas: list, month: str) -> str:
    wb = Workbook()
    ws = wb.active

    year, mon = month.split("-")
    month_name = MONTH_NAMES_AZ.get(mon, mon)
    ws.title = f"{month_name} {year}"

    # Sütun genişlikləri
    for col, width in zip("ABCDEF", [6, 30, 16, 16, 16, 16]):
        ws.column_dimensions[col].width = width

    row = 1
    # Başlıq
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, f"{COMPANY_NAME} — {month_name} {year} Hesabatı")
    c.font = _font(bold=True, size=14, color=WHITE)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 28
    row += 1

    # Xülasə bölməsi
    summary_data = [
        ("Yeni smetalar",          report["new_smetas"],           None),
        ("Tamamlanan",             report["completed"],            None),
        ("Aktiv layihələr",        report["active"],               None),
        ("Ümumi smeta dəyəri",     f"{report['total_value']:,.2f} AZN", None),
        ("Alınan ödənişlər",       f"{report['received_payments']:,.2f} AZN", None),
        ("Gözlənilən ödənişlər",   f"{report['expected_payments']:,.2f} AZN", None),
        ("Material xərcləri",      f"{report['material_costs']:,.2f} AZN", None),
        ("Mənfəət (təxmini)",      f"{report['profit']:,.2f} AZN", None),
    ]

    for label, value, _ in summary_data:
        ws.merge_cells(f"A{row}:C{row}")
        lc = ws.cell(row, 1, label)
        lc.font = _font(size=10)
        lc.fill = _fill(LIGHT_GRAY)
        lc.alignment = _align("left")
        lc.border = _border()

        ws.merge_cells(f"D{row}:F{row}")
        vc = ws.cell(row, 4, value)
        vc.font = _font(bold=True, size=10)
        vc.alignment = _align("right")
        vc.border = _border()
        row += 1

    row += 1

    # Smeta cədvəli
    headers = ["№", "Smeta nömrəsi", "Müştəri", "Ümumi", "Ödənilib", "Status"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row, col_idx, h)
        c.font = _font(bold=True, size=9, color=WHITE)
        c.fill = _fill(DARK_BLUE)
        c.alignment = _align("center")
        c.border = _border()
    ws.row_dimensions[row].height = 20
    row += 1

    STATUS_MAP = {
        "draft":    "Qaralama",
        "sent":     "Göndərilib",
        "approved": "Təsdiqlənib",
        "rejected": "Rədd edilib",
        "active":   "Aktiv",
    }

    for idx, s in enumerate(smetas, 1):
        row_data = [
            idx,
            s["smeta_number"],
            s["client_name"],
            f"{s['total']:,.2f}",
            f"{s.get('paid', 0):,.2f}",
            STATUS_MAP.get(s["status"], s["status"]),
        ]
        fill = _fill(LIGHT_GRAY) if idx % 2 == 0 else _fill(WHITE)
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row, col_idx, val)
            c.font = _font(size=9)
            c.fill = fill
            c.alignment = _align("center" if col_idx in (1, 6) else "left")
            c.border = _border()
        row += 1

    fname = os.path.join(OUTPUT_DIR, f"hesabat_{month}.xlsx")
    wb.save(fname)
    return fname


# ═══════════════════════════════════════════════════════════════════════════════
#  MÜQAVİLƏ PDF
# ═══════════════════════════════════════════════════════════════════════════════

def generate_contract_pdf(smeta: dict) -> str:
    fname = os.path.join(OUTPUT_DIR, f"muqavile_{smeta['smeta_number']}.pdf")
    doc = SimpleDocTemplate(
        fname, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    dark = colors.HexColor(f"#{DARK_BLUE}")
    gold = colors.HexColor(f"#{ACCENT}")

    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    story = []

    # Başlıq
    story.append(Paragraph(
        f"<b>MÜQAVİLƏ</b>",
        ps("Title", fontName=FONT_BOLD, fontSize=18, textColor=dark,
           alignment=TA_CENTER, spaceAfter=6)
    ))
    story.append(Paragraph(
        f"№ {smeta['smeta_number']}",
        ps("SubTitle", fontName=FONT, fontSize=11, textColor=gold,
           alignment=TA_CENTER, spaceAfter=4)
    ))
    story.append(Paragraph(
        f"Bakı şəhəri, {smeta['created_at'][:10]}",
        ps("Date", fontName=FONT, fontSize=9, textColor=colors.grey,
           alignment=TA_CENTER, spaceAfter=20)
    ))

    story.append(HRFlowable(width="100%", thickness=2, color=dark))
    story.append(Spacer(1, 0.5*cm))

    # Tərəflər
    story.append(Paragraph(
        "<b>TƏRƏFLƏRİN REKVİZİTLƏRİ</b>",
        ps("H2", fontName=FONT_BOLD, fontSize=11, textColor=dark, spaceAfter=8)
    ))

    parties_data = [
        ["İCRAÇI:",    COMPANY_NAME],
        ["Ünvan:",      COMPANY_ADDRESS],
        ["Telefon:",    COMPANY_PHONE],
        ["",           ""],
        ["SİFARİŞÇİ:", smeta["client_name"]],
        ["Telefon:",    smeta["client_phone"]],
        ["Ünvan:",      smeta["address"]],
    ]
    for label, value in parties_data:
        if not label and not value:
            story.append(Spacer(1, 0.2*cm))
            continue
        story.append(Paragraph(
            f"<b>{label}</b> {value}",
            ps("P", fontName=FONT, fontSize=10, spaceAfter=3)
        ))

    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
    story.append(Spacer(1, 0.5*cm))

    # Mövzu
    story.append(Paragraph(
        "<b>1. MÜQAVİLƏNİN MƏVZUSu</b>",
        ps("H2", fontName=FONT_BOLD, fontSize=11, textColor=dark, spaceAfter=8)
    ))
    rooms_list = list(smeta["rooms_data"].keys()) if smeta.get("rooms_data") else []
    rooms_txt = ", ".join(rooms_list) if rooms_list else "—"
    story.append(Paragraph(
        f"İcraçı Sifarişçinin aşağıdaki obyektdə təmir-bərpa işlərini yerinə yetirməyi öhdəsinə götürür:\n"
        f"<b>Ünvan:</b> {smeta['address']}\n"
        f"<b>Otaqlar:</b> {rooms_txt}",
        ps("P", fontName=FONT, fontSize=10, spaceAfter=8, leading=16)
    ))

    # Dəyər
    story.append(Paragraph(
        "<b>2. MÜQAVİLƏNİN DƏYƏRİ VƏ ÖDƏNİŞ ŞƏRTI</b>",
        ps("H2", fontName=FONT_BOLD, fontSize=11, textColor=dark, spaceAfter=8)
    ))
    story.append(Paragraph(
        f"İşlərin ümumi dəyəri: <b>{smeta['total']:,.2f} AZN</b> (ƏDV daxil deyil)\n\n"
        f"Ödəniş cədvəli:\n"
        f"• Müqavilə imzalandıqdan sonra avans — 30%\n"
        f"• İşlər 50% tamamlandıqda — 40%\n"
        f"• İşlər tamamlandıqda son ödəniş — 30%",
        ps("P", fontName=FONT, fontSize=10, spaceAfter=8, leading=16)
    ))

    # Müddət
    story.append(Paragraph(
        "<b>3. İŞLƏRİN MÜDDƏTİ</b>",
        ps("H2", fontName=FONT_BOLD, fontSize=11, textColor=dark, spaceAfter=8)
    ))
    story.append(Paragraph(
        "İşlər müqavilə imzalandıqdan sonra razılaşdırılmış müddətdə tamamlanacaq.\n"
        "Başlama tarixi: _______________     Bitmə tarixi: _______________",
        ps("P", fontName=FONT, fontSize=10, spaceAfter=8, leading=16)
    ))

    # Zəmanət
    story.append(Paragraph(
        "<b>4. ZƏMANƏT</b>",
        ps("H2", fontName=FONT_BOLD, fontSize=11, textColor=dark, spaceAfter=8)
    ))
    story.append(Paragraph(
        "İcraçı yerinə yetirilən işlərə <b>2 (iki) il</b> zəmanət verir.",
        ps("P", fontName=FONT, fontSize=10, spaceAfter=8)
    ))

    # Ümumi müddəalar
    story.append(Paragraph(
        "<b>5. ÜMUMI MÜDDƏALAR</b>",
        ps("H2", fontName=FONT_BOLD, fontSize=11, textColor=dark, spaceAfter=8)
    ))
    general = [
        "Materiallar ayrıca hesablanır və müştəri tərəfindən razılaşdırılır.",
        "İşlər müvafiq standartlara uyğun yerinə yetirilir.",
        "Müqavilə iki nüsxədə imzalanır, hər bir tərəfdə bir nüsxə saxlanılır.",
    ]
    for item in general:
        story.append(Paragraph(
            f"• {item}",
            ps("Li", fontName=FONT, fontSize=10, spaceAfter=4, leftIndent=12)
        ))

    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
    story.append(Spacer(1, 0.5*cm))

    # İmzalar
    story.append(Paragraph(
        "<b>TƏRƏFLƏRİN İMZALARI</b>",
        ps("H2", fontName=FONT_BOLD, fontSize=11, textColor=dark, spaceAfter=12,
           alignment=TA_CENTER)
    ))

    sign_tbl = Table(
        [
            ["İCRAÇI",                        "SİFARİŞÇİ"],
            [COMPANY_NAME,                    smeta["client_name"]],
            [f"Tel: {COMPANY_PHONE}",         f"Tel: {smeta['client_phone']}"],
            ["", ""],
            ["İmza: ___________________",    "İmza: ___________________"],
            ["Tarix: ___________________",   "Tarix: ___________________"],
        ],
        colWidths=[9*cm, 9*cm]
    )
    sign_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0, 0), (-1, -1), FONT),
        ("FONTNAME",   (0, 0), (-1, 0),  FONT_BOLD),
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("TEXTCOLOR",  (0, 0), (-1, 0),  dark),
        ("ALIGN",      (0, 0), (0, -1),  "LEFT"),
        ("ALIGN",      (1, 0), (1, -1),  "LEFT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(sign_tbl)

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"{COMPANY_NAME} | {COMPANY_PHONE} | {COMPANY_ADDRESS}",
        ps("Footer", fontName=FONT, fontSize=8, textColor=colors.grey,
           alignment=TA_CENTER)
    ))

    doc.build(story)
    return fname
